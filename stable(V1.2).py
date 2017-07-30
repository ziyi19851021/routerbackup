#!/usr/bin/env python
# coding=gbk

import datetime
import time
import sys
import paramiko
import os
from openpyxl import Workbook
from openpyxl import load_workbook

global static1, static2 

def check_ip(ipaddr, device, username, old_password, new_password, log):
	global static1, static2

	addr=ipaddr.strip().split('.')  #�и�IP��ַΪһ���б�
	#print addr
	if len(addr) != 4:  #�и���б������4������
		content = ipaddr + u"������ip�淶!\n"
		static2 = static2 + 1
		print content
		log.write(content.encode("gbk"))
		return 0
	else:
		try:
			for i in range(4):
				addr[i]=int(addr[i])  #ÿ����������Ϊ���֣�����У��ʧ��
		except:
			content = ipaddr + u"������ip�淶!\n"
			static2 = static2 + 1
			print content
			log.write(content.encode("gbk"))
			return 0

		if addr[0] == 11 :
			if addr[1] == 76 or addr[1] == 77 :
				if addr[2] >= 0 and addr[2] <= 255:
					if addr[3] >=0 and addr[3] <=255:
						pass
					else:
						content = ipaddr + u"������ip�淶!\n"
						static2 = static2 + 1
						print content
						log.write(content.encode("gbk"))
						return 0
				else:
					content = ipaddr + u"������ip�淶!\n"
					static2 = static2 + 1
					print content
					log.write(content.encode("gbk"))
					return 0
			else:
				content = ipaddr + u"������ip�淶!\n"
				static2 = static2 + 1
				print content
				log.write(content.encode("gbk"))
				return 0
		else:
			content = ipaddr + u"������ip�淶!\n"
			static2 = static2 + 1
			print content
			log.write(content.encode("gbk"))
			return 0

	if device == None:
		content = ipaddr + u"�������豸���ƹ淶!\n"
		static2 = static2 + 1
		print content
		log.write(content.encode("gbk"))
		return 0
	if username == None:
		content = ipaddr + u"�������û����淶!\n"
		static2 = static2 + 1
		print content
		log.write(content.encode("gbk"))
		return 0
	if old_password == None:
		content = ipaddr + u"�����Ͼ�����淶!\n"
		static2 = static2 + 1
		print content
		log.write(content.encode("gbk"))
		return 0
	if new_password == None:
		content = ipaddr + u"������������淶!\n"
		static2 = static2 + 1
		print content
		log.write(content.encode("gbk"))
		return 0
	return 2

def cisco_disable_paging(remote_conn):
	'''Disable paging on a Cisco router'''

	remote_conn.send("terminal length 0\n")
	time.sleep(1)

	# Clear the buffer on the screen
	output = remote_conn.recv(1000)
	return output

def h3c_disable_paging(remote_conn):
	remote_conn.send("screen-length 0\n")
	time.sleep(1)

	output = remote_conn.recv(1000)
	return output


def do_cisco(ip, username, password, newpassword, log):
	global static1, static2
	remote_conn_pre = paramiko.SSHClient()

	remote_conn_pre.set_missing_host_key_policy(paramiko.AutoAddPolicy())
	try:
		remote_conn_pre.connect(ip, username=username, password=password, look_for_keys=False, allow_agent=False)

		remote_conn = remote_conn_pre.invoke_shell()
    
		remote_conn.send("\n")
		remote_conn.send("en\n")
		time.sleep(1)
		remote_conn.send(password + "\n")
		remote_conn.send("config terminal\n")
		remote_conn.send("enable secret "+newpassword+"\n")
		time.sleep(1)
		remote_conn.send("username " + username + " password "+ newpassword + "\n")
		time.sleep(1)
		remote_conn.send("end\n")
		time.sleep(1)
		remote_conn.send("write\n")
		time.sleep(1)
		static1 = static1 + 1
		content = ip + u"���IP���ڵ��豸�����Ѿ��ɹ��޸�!\n"
		print content
		log.write(content.encode("gbk"))

	except:
		static2 = static2 + 1
		content = ip+ u"���IP�����ڻ����û��������д���!\n"
		print content
		log.write(content.encode("gbk"))

def do_h3c(ip, username, password, newpassword, log):
	global static1, static2
	remote_conn_pre = paramiko.SSHClient()

	remote_conn_pre.set_missing_host_key_policy(paramiko.AutoAddPolicy())
	try:
		remote_conn_pre.connect(ip, username=username, password=password, look_for_keys=False, allow_agent=False)

		remote_conn = remote_conn_pre.invoke_shell()

		remote_conn.send("\n")
		remote_conn.send("sys\n")
		remote_conn.send("local-user " + username + "\n")
		remote_conn.send("password simple "+newpassword+"\n")
		time.sleep(1)
		remote_conn.send("quit\n")
		time.sleep(1)
		remote_conn.send("save\n")
		time.sleep(1)
		remote_conn.send("y\n")
		time.sleep(1)
		remote_conn.send("\n")
		time.sleep(1)
		remote_conn.send("y\n")
		time.sleep(1)
		static1 = static1 + 1
		content = ip + u"���IP���ڵ��豸�����Ѿ����ɹ��޸�!\n"
		print content
		log.write(content.encode("gbk"))
	except:
		static2 = static2 + 1
		content = ip+ u"���IP�����ڻ����û��������д���!\n"
		print content
		log.write(content.encode("gbk"))

def display():
	global static1, static2

	control = 0
	while control < 3:
		print "����������ѡ����س������в�����"
		print "R���������������б��е��豸����"
		print "V�����鿴�汾��Ϣ����������"
		print "G��������һ�������ļ�ģ��"
		print "Q�����˳�������"
		v = raw_input("Input:")
		if v == "r" or v == "R":
			ti = time.strftime('%Y%m%d%H%M%S', time.localtime(time.time()))	
			log = open(ti + '_log.txt', "w")
			content = time.strftime('%Y��%m��%d��%Hʱ%M��%S��', time.localtime(time.time())) + "\n"	
			log.write(content)
			static1 = 0
			static2 = 0
			control = 1
			do_getInfo(log)
			static = static1 + static2
			content = u"�����޸Ĺ��漰�豸"+str(static)+u"̨,���гɹ��޸�"+str(static1)+u"̨���޸�ʧ��"+str(static2)+u"̨��"
			print content
			log.write(content.encode("gbk"))
			log.close()
		elif v == "v" or v == "V": 
			control = 2 
			view_info()
		elif v == "q" or v == "Q":
			control = 3 
		elif v == "g" or v == "G":
			do_saveinfo()
		else:
			print "�����������������!!!"

def view_info():
	print "                          <��������>"
	print "    ������Ϊ������֧�Ƽ��ƿ����������޸������豸���빤�ߣ���ʽ��V1.2����������ʹ�ñ����ߣ�������ͬ���������"
	print "    1���򱾳���©����bug����ɵ�һ����ʧ��������֧�Ƽ����޹ء�"
	print "    2������������κ��йر������©����bug,�뼰ʱ��ϵ������֧�Ƽ��ƣ���ϵ�绰��0851-28220393��"

def do_getInfo(log):
	global static1, static2 
	
	i=1
	st =True
	try:
		wb = load_workbook(filename = 'Password.xlsx')
		sheet_ranges = wb['Sheet']
		while st:
			i= i + 1
			ip = sheet_ranges['A' + str(i)].value
			if ip == None:
				break
			else:
				ip = ip.strip()
				device = sheet_ranges['C' + str(i)].value
				username = sheet_ranges['D' + str(i)].value
				old_password = sheet_ranges['E' + str(i)].value
				new_password = sheet_ranges['F' + str(i)].value
				if check_ip(ip, device, username, old_password, new_password, log) == 2:
					device = device.strip()
					username = username.strip()
					old_password = old_password.strip()
					new_password = new_password.strip()
					if device == 'cisco' or device == 'Cisco' or device == 'CISCO': 
						do_cisco(ip, username, old_password, new_password, log)
					elif device == 'h3c' or device == 'H3C':
						do_h3c(ip, username, old_password, new_password, log)
					elif device == 'mp' or device == 'MP' or device == 'Mp':
						do_cisco(ip, username, old_password, new_password, log)
					else :
						content = ip + u"���IP���������豸��Ϣ���ڴ���!\n"
						static2 = static2 + 1
						print content
						log.write(content.encode("gbk"))
					
	except:
		content = "Password�ļ������ڣ�\n"
		print content

def do_saveinfo():
	try:
		wb = load_workbook(filename = 'Password.xlsx')
		print "Password�ļ��Ѿ�����!!"
	except:
		wb = Workbook()

	# grab the active worksheet
		ws = wb.active

	# Data can be assigned directly to cells
	#ws['A1'] = 'IP Adress'

	# Rows can also be appended
		ws.append([u'IP��ַ', u"���ӷ�ʽ", u"�豸����", u"�û���", u"������", u"������"])

	# Python types will automatically be converted
	#ws['A2'] = datetime.datetime.now()

	# Save the file
		wb.save("Password.xlsx")

if __name__ == '__main__':
	view_info()
	#do_getinfo()
	display()

